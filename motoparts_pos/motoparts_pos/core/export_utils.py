"""
Export helpers: CSV always works (stdlib). Excel needs openpyxl, PDF needs reportlab.
Both degrade gracefully with a clear message if the optional library is missing.
"""

import csv
import os


def export_to_csv(filepath, headers, rows):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
    return filepath


def export_to_excel(filepath, headers, rows, title="Report"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise RuntimeError(
            "Excel export requires the 'openpyxl' package. Install it with: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Report"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows] or [10])
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_len + 4, 40)

    wb.save(filepath)
    return filepath


def export_to_pdf(filepath, title, headers, rows, subtitle=""):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError as e:
        raise RuntimeError(
            "PDF export requires the 'reportlab' package. Install it with: pip install reportlab"
        ) from e

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                             topMargin=15 * mm, bottomMargin=15 * mm)
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [headers] + [[str(c) for c in r] for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    return filepath


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path
